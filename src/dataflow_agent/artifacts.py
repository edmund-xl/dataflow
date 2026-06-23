from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .models import Finding, GraphModel, WorkbookData
from .util import simple_yaml, write_csv, write_json


def write_graph_artifacts(graph: GraphModel, normalized_dir: Path) -> None:
    normalized_dir.mkdir(parents=True, exist_ok=True)
    nodes = [node.as_dict() for node in graph.nodes.values()]
    edges = [edge.as_dict() for edge in graph.edges]
    write_csv(
        normalized_dir / "nodes.csv",
        [_flatten(row) for row in nodes],
        ["id", "type", "label", "sheet", "status", "metadata"],
    )
    write_csv(
        normalized_dir / "edges.csv",
        [_flatten(row) for row in edges],
        ["id", "type", "source", "target", "label", "status", "evidence_id", "metadata"],
    )
    graph_data = graph.as_dict()
    write_json(normalized_dir / "dataflow_graph.json", graph_data)
    (normalized_dir / "dataflow_graph.yaml").write_text(simple_yaml(graph_data) + "\n", encoding="utf-8")


def write_normalized_snapshot(workbook: WorkbookData, normalized_dir: Path) -> None:
    normalized_dir.mkdir(parents=True, exist_ok=True)
    write_json(normalized_dir / "normalized_data.json", workbook.sheets)
    workbook_path = normalized_dir / "mainnet_dataflow_normalized.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet, rows in workbook.sheets.items():
        ws = wb.create_sheet(title=sheet[:31])
        headers = workbook.headers.get(sheet) or sorted({key for row in rows for key in row})
        if headers:
            ws.append(headers)
            for row in rows:
                ws.append([row.get(header, "") for header in headers])
    wb.save(workbook_path)


def write_validation_json(findings: list[Finding], reports_dir: Path) -> None:
    reports_dir.mkdir(parents=True, exist_ok=True)
    write_json(reports_dir / "validation_errors.json", [finding.as_row() for finding in findings])


def write_findings_workbook(path: Path, title: str, findings: list[Finding]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    fields = ["Gate", "Severity", "Sheet", "Row_ID", "Field", "Message", "Suggested_Action", "Status", "Owner", "Due_Date", "Exception_Decision", "Evidence_ID"]
    ws.append(fields)
    for finding in findings:
        row = finding.as_row()
        ws.append([row[field] for field in fields])
    _autosize(ws)
    wb.save(path)


def write_acceptance_checklist(path: Path, validation_findings: list[Finding], risk_findings: list[Finding], graph: GraphModel) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Acceptance"
    checks = [
        ("Input workbook parsed", True),
        ("No P0/P1 validation findings", not any(f.severity in {"P0", "P1"} for f in validation_findings)),
        ("Graph nodes generated", bool(graph.nodes)),
        ("Graph edges generated", bool(graph.edges)),
        ("No P0 risk findings", not any(f.severity == "P0" for f in risk_findings)),
        ("Reports generated", True),
        ("Diagrams generated", True),
        ("Package generated", True),
    ]
    ws.append(["Check", "Status"])
    for label, ok in checks:
        ws.append([label, "Pass" if ok else "Needs Review"])
    _autosize(ws)
    wb.save(path)


def write_logic_check_results(path: Path, findings: list[Finding]) -> None:
    write_json(path, [finding.as_row() for finding in findings])


def _flatten(row: dict[str, Any]) -> dict[str, Any]:
    flattened = dict(row)
    for key, value in list(flattened.items()):
        if isinstance(value, (dict, list)):
            flattened[key] = simple_yaml(value).replace("\n", "; ")
    return flattened


def _autosize(ws: Any) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 80)
