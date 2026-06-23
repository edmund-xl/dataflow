from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .constants import find_workbook
from .models import WorkbookData
from .schema import load_schema
from .util import write_json
from .xlsx_reader import read_workbook


MERGED_WORKBOOK_NAME = "dataflow_collection_template_v0.1.xlsx"


@dataclass
class MergeIssue:
    issue_type: str
    severity: str
    sheet: str
    primary_key: str
    value: str
    source: str
    message: str

    def as_dict(self) -> dict[str, str]:
        return {
            "issue_type": self.issue_type,
            "severity": self.severity,
            "sheet": self.sheet,
            "primary_key": self.primary_key,
            "value": self.value,
            "source": self.source,
            "message": self.message,
        }


@dataclass
class MergeResult:
    merged_dcp: Path
    workbook_path: Path
    issues: list[MergeIssue] = field(default_factory=list)
    source_count: int = 0
    row_count: int = 0
    duplicate_count: int = 0
    conflict_count: int = 0
    lineage: list[dict[str, Any]] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "merged_dcp": str(self.merged_dcp),
            "workbook_path": str(self.workbook_path),
            "source_count": self.source_count,
            "row_count": self.row_count,
            "duplicate_count": self.duplicate_count,
            "conflict_count": self.conflict_count,
            "lineage": self.lineage,
            "issues": [issue.as_dict() for issue in self.issues],
        }


def merge_dcps(inputs: list[Path], output_root: Path, version: str) -> MergeResult:
    if len(inputs) < 1:
        raise ValueError("At least one DCP is required.")
    schema = load_schema()
    merged_dcp = output_root / f"merged_dcp_{version}"
    merged_dcp.mkdir(parents=True, exist_ok=True)
    workbooks = [_read_source(path, schema) for path in inputs]
    merged_sheets, headers, issues, duplicate_count, conflict_count, lineage = _merge_workbooks(workbooks, inputs, schema)
    workbook_path = merged_dcp / MERGED_WORKBOOK_NAME
    _write_workbook(workbook_path, merged_sheets, headers, schema)
    result = MergeResult(
        merged_dcp=merged_dcp,
        workbook_path=workbook_path,
        issues=issues,
        source_count=len(inputs),
        row_count=sum(len(rows) for rows in merged_sheets.values()),
        duplicate_count=duplicate_count,
        conflict_count=conflict_count,
        lineage=lineage,
    )
    _write_merge_reports(merged_dcp, result)
    return result


def _read_source(path: Path, schema: dict) -> WorkbookData:
    workbook_path = find_workbook(path.resolve())
    return read_workbook(workbook_path, schema)


def _merge_workbooks(
    workbooks: list[WorkbookData],
    sources: list[Path],
    schema: dict,
) -> tuple[dict[str, list[dict[str, str]]], dict[str, list[str]], list[MergeIssue], int, int, list[dict[str, Any]]]:
    merged: dict[str, list[dict[str, str]]] = {}
    headers: dict[str, list[str]] = {}
    issues: list[MergeIssue] = []
    duplicate_count = 0
    conflict_count = 0
    lineage_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}

    for sheet in schema["required_sheets"]:
        pk = schema["sheets"][sheet].get("primary_key") or ""
        headers[sheet] = schema["sheets"][sheet]["fields"]
        by_key: dict[str, dict[str, str]] = {}
        rows_without_key: list[dict[str, str]] = []
        for workbook, source in zip(workbooks, sources, strict=True):
            for row in workbook.sheets.get(sheet, []):
                normalized = {field: row.get(field, "") for field in headers[sheet]}
                key = _row_key(sheet, pk, normalized)
                if not key:
                    rows_without_key.append(normalized)
                    continue
                if key not in by_key:
                    by_key[key] = normalized
                    lineage_by_key[(sheet, pk, key)] = {"sheet": sheet, "primary_key": pk, "value": key, "kept_source": str(source), "sources": [str(source)]}
                    continue
                lineage_by_key[(sheet, pk, key)]["sources"].append(str(source))
                if _same_row(by_key[key], normalized):
                    duplicate_count += 1
                    issues.append(MergeIssue("Duplicate", "Info", sheet, pk, key, str(source), "Duplicate row is identical and was de-duplicated."))
                    continue
                conflict_count += 1
                issues.append(MergeIssue("Conflict", "P1", sheet, pk, key, str(source), "Same primary key has different values; first row was kept."))
        deduped_keyless = _dedupe_rows_without_key(sheet, rows_without_key, issues, lineage_by_key)
        merged[sheet] = list(by_key.values()) + deduped_keyless
    return merged, headers, issues, duplicate_count, conflict_count, list(lineage_by_key.values())


def _row_key(sheet: str, pk: str, row: dict[str, str]) -> str:
    if pk:
        return row.get(pk, "")
    if sheet == "90_Enums":
        return f"{row.get('Enum_Name', '')}:{row.get('Value', '')}"
    return ""


def _same_row(a: dict[str, str], b: dict[str, str]) -> bool:
    keys = set(a).union(b)
    return all((a.get(key, "") or "") == (b.get(key, "") or "") for key in keys)


def _dedupe_rows_without_key(sheet: str, rows: list[dict[str, str]], issues: list[MergeIssue], lineage_by_key: dict[tuple[str, str, str], dict[str, Any]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        key = json.dumps(row, ensure_ascii=False, sort_keys=True)
        if key in seen:
            issues.append(MergeIssue("Duplicate", "Info", sheet, "", "", "", "Duplicate keyless row was de-duplicated."))
            continue
        seen.add(key)
        lineage_by_key[(sheet, "", key)] = {"sheet": sheet, "primary_key": "", "value": key, "kept_source": "", "sources": [""]}
        result.append(row)
    return result


def _write_workbook(path: Path, sheets: dict[str, list[dict[str, str]]], headers: dict[str, list[str]], schema: dict) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet in schema["required_sheets"]:
        ws = wb.create_sheet(sheet[:31])
        sheet_headers = headers[sheet]
        ws.append(sheet_headers)
        for row in sheets.get(sheet, []):
            ws.append([row.get(header, "") for header in sheet_headers])
    wb.save(path)


def _write_merge_reports(merged_dcp: Path, result: MergeResult) -> None:
    write_json(merged_dcp / "merge_report.json", result.as_dict())
    write_json(merged_dcp / "merge_lineage.json", result.lineage)
    wb = Workbook()
    ws = wb.active
    ws.title = "Merge_Report"
    fields = ["issue_type", "severity", "sheet", "primary_key", "value", "source", "message"]
    ws.append(fields)
    for issue in result.issues:
        row = issue.as_dict()
        ws.append([row[field] for field in fields])
    wb.save(merged_dcp / "merge_report.xlsx")
