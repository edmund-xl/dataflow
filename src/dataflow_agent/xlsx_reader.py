from __future__ import annotations

import re
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from .models import Row, WorkbookData
from .util import clean

NS = {
    "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def read_workbook(path: Path, schema: dict) -> WorkbookData:
    sheets, headers = _read_with_openpyxl(path, schema)
    if _reader_failed(sheets, headers, schema):
        sheets, headers = _read_with_xml(path, schema)
    metadata = _metadata(sheets.get("00_Metadata", []))
    enums = _enums(sheets.get("90_Enums", []), schema)
    return WorkbookData(path=path, sheets=sheets, headers=headers, metadata=metadata, enums=enums)


def _reader_failed(sheets: dict[str, list[Row]], headers: dict[str, list[str]], schema: dict) -> bool:
    required = schema["required_sheets"]
    if not all(sheet in sheets for sheet in required):
        return True
    for sheet in required:
        expected = schema["sheets"][sheet]["fields"]
        if not headers.get(sheet) or expected[0] not in headers[sheet]:
            return True
    return False


def _read_with_openpyxl(path: Path, schema: dict) -> tuple[dict[str, list[Row]], dict[str, list[str]]]:
    try:
        import openpyxl
    except Exception:
        return {}, {}

    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return {}, {}

    sheets: dict[str, list[Row]] = {}
    headers: dict[str, list[str]] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = []
        for raw in ws.iter_rows(values_only=True):
            row = [clean(value) for value in raw]
            if any(row):
                rows.append(row)
        parsed_header, data_rows = _parse_rows(sheet_name, rows, schema)
        headers[sheet_name] = parsed_header
        sheets[sheet_name] = data_rows
    return sheets, headers


def _read_with_xml(path: Path, schema: dict) -> tuple[dict[str, list[Row]], dict[str, list[str]]]:
    with ZipFile(path) as zf:
        shared_strings = _shared_strings(zf)
        sheet_paths = _sheet_paths(zf)
        sheets: dict[str, list[Row]] = {}
        headers: dict[str, list[str]] = {}
        for sheet_name, xml_path in sheet_paths:
            rows = _xml_rows(zf, xml_path, shared_strings)
            parsed_header, data_rows = _parse_rows(sheet_name, rows, schema)
            headers[sheet_name] = parsed_header
            sheets[sheet_name] = data_rows
        return sheets, headers


def _sheet_paths(zf: ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels
        if rel.attrib.get("Target", "").startswith("worksheets/")
    }
    results = []
    for sheet in workbook.findall(".//x:sheet", NS):
        rel_id = sheet.attrib[f"{{{NS['r']}}}id"]
        target = rel_map[rel_id]
        results.append((sheet.attrib["name"], f"xl/{target}"))
    return results


def _shared_strings(zf: ZipFile) -> list[str]:
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    strings = []
    for item in root.findall("x:si", NS):
        texts = [node.text or "" for node in item.findall(".//x:t", NS)]
        strings.append("".join(texts))
    return strings


def _xml_rows(zf: ZipFile, xml_path: str, shared_strings: list[str]) -> list[list[str]]:
    root = ET.fromstring(zf.read(xml_path))
    rows: list[list[str]] = []
    for row in root.findall(".//x:sheetData/x:row", NS):
        values: dict[int, str] = {}
        for cell in row.findall("x:c", NS):
            ref = cell.attrib.get("r", "A0")
            value_node = cell.find("x:v", NS)
            inline = cell.find("x:is/x:t", NS)
            value = ""
            if inline is not None:
                value = inline.text or ""
            elif value_node is not None:
                value = value_node.text or ""
                if cell.attrib.get("t") == "s" and value.isdigit():
                    idx = int(value)
                    value = shared_strings[idx] if idx < len(shared_strings) else value
            values[_colnum(ref)] = clean(value)
        if values:
            rows.append([values.get(idx, "") for idx in range(1, max(values) + 1)])
    return rows


def _colnum(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    if not match:
        return 1
    value = 0
    for ch in match.group(1):
        value = value * 26 + ord(ch) - 64
    return value


def _parse_rows(sheet_name: str, rows: list[list[str]], schema: dict) -> tuple[list[str], list[Row]]:
    if sheet_name not in schema["sheets"]:
        return [], []
    expected = schema["sheets"][sheet_name]["fields"]
    header_idx = _find_header_row(rows, expected)
    if header_idx is None:
        return [], []
    header = _trim_blank_tail(rows[header_idx])
    header_len = len(header)
    data: list[Row] = []
    for raw in rows[header_idx + 1 :]:
        padded = raw + [""] * max(0, header_len - len(raw))
        row = {header[i]: clean(padded[i]) for i in range(header_len) if header[i]}
        if any(row.values()):
            data.append(row)
    return header, data


def _find_header_row(rows: list[list[str]], expected: list[str]) -> int | None:
    expected_set = set(expected)
    best_idx: int | None = None
    best_score = 0
    for idx, row in enumerate(rows):
        values = [cell for cell in row if cell]
        score = len(expected_set.intersection(values))
        if score > best_score:
            best_score = score
            best_idx = idx
    if best_score >= min(3, len(expected)):
        return best_idx
    return None


def _trim_blank_tail(row: list[str]) -> list[str]:
    values = list(row)
    while values and not values[-1]:
        values.pop()
    return values


def _metadata(rows: list[Row]) -> dict[str, str]:
    result = {}
    for row in rows:
        field = row.get("Field", "")
        if field:
            result[field] = row.get("Value", "")
    return result


def _enums(rows: list[Row], schema: dict) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {key: list(values) for key, values in schema.get("enum_fields", {}).items()}
    for row in rows:
        name = row.get("Enum_Name", "")
        value = row.get("Value", "")
        if name and value and value not in result.setdefault(name, []):
            result[name].append(value)
    return result

