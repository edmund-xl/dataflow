from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

from dataflow_agent.analyzers.change_diff import build_change_diff, write_change_diff_report
from dataflow_agent.cli import main


def test_change_diff_detects_workbook_risks_from_dcp_sources(tmp_path: Path) -> None:
    base_dir = tmp_path / "base"
    new_dir = tmp_path / "new"
    shutil.copytree("samples/DCP_clean_v0.1", base_dir)
    shutil.copytree("samples/DCP_clean_v0.1", new_dir)
    workbook_path = new_dir / "dataflow_collection_template_v0.1.xlsx"
    wb = load_workbook(workbook_path)
    _remove_first_data_row(wb["10_Monitoring"])
    _append_row(
        wb["04_Services"],
        {
            "Record_ID": "rec-svc-public-entry-new",
            "Service_ID": "svc-public-entry-new",
            "Service_Name": "Public Entry New",
            "Service_Priority": "P0",
            "Running_On_Instance_ID": "inst-web-1",
            "Protocol": "TCP",
            "Listen_Ports": "443",
            "Service_Owner": "team-entry",
            "Service_Role": "public entry",
            "Environment": "testnetv2",
            "Evidence_ID": "ev-svc-public-entry-new",
            "Confirmation_Status": "Confirmed",
        },
    )
    _append_row(
        wb["12_External_Services"],
        {
            "Record_ID": "rec-ext-new",
            "External_ID": "ext-new",
            "External_Name": "New External API",
            "Endpoint": "https://api.example.invalid",
            "Protocol": "HTTPS",
            "Port": "443",
            "Used_By_Service_ID": "svc-public-entry-new",
            "Environment": "testnetv2",
            "Evidence_ID": "ev-ext-new",
            "Confirmation_Status": "Confirmed",
        },
    )
    _append_row(
        wb["06_Data_Assets"],
        {
            "Record_ID": "rec-data-new-secret",
            "Data_Asset_ID": "data-new-secret",
            "Data_Asset_Name": "New Secret Store",
            "Data_Asset_Type": "database",
            "Project_ID": "proj-demo",
            "Used_By_Service_ID": "svc-public-entry-new",
            "Access_Type": "read",
            "Sensitivity": "High",
            "Environment": "testnetv2",
            "Evidence_ID": "ev-data-new-secret",
            "Confirmation_Status": "Confirmed",
        },
    )
    _append_row(
        wb["09_IAM_SA"],
        {
            "Record_ID": "rec-iam-new-owner",
            "IAM_Binding_ID": "iam-new-owner",
            "Service_Account_ID": "sa-new-owner",
            "Service_Account_Email": "sa-new-owner@example.invalid",
            "Used_By_Service_ID": "svc-public-entry-new",
            "Role": "roles/owner",
            "Scope": "project",
            "Is_High_Privilege": "Yes",
            "Environment": "testnetv2",
            "Evidence_ID": "ev-iam-new-owner",
            "Confirmation_Status": "Confirmed",
        },
    )
    wb.save(workbook_path)

    diff = build_change_diff(base_dir, new_dir)

    categories = {risk["category"] for risk in diff["risks"]}
    assert {
        "added_public_service",
        "added_external_call",
        "added_sensitive_data_access",
        "added_high_privilege_iam",
        "removed_monitoring",
    }.issubset(categories)
    assert diff["summary"]["added"] > 0
    assert diff["summary"]["removed"] > 0


def test_change_diff_supports_graph_json_and_writes_reports(tmp_path: Path) -> None:
    base_graph = tmp_path / "base_graph.json"
    new_graph = tmp_path / "new_graph.json"
    base_graph.write_text(
        json.dumps(
            {
                "nodes": [
                    {"id": "svc-api", "type": "service", "label": "API", "sheet": "04_Services"},
                    {"id": "ext-payments", "type": "external_service", "label": "Payments", "sheet": "12_External_Services"},
                ],
                "edges": [],
                "dropped_edges": [],
            }
        ),
        encoding="utf-8",
    )
    new_graph.write_text(
        json.dumps(
            {
                "nodes": [
                    {"id": "svc-api", "type": "service", "label": "API", "sheet": "04_Services"},
                    {"id": "ext-payments", "type": "external_service", "label": "Payments", "sheet": "12_External_Services"},
                ],
                "edges": [
                    {
                        "id": "edge-ext",
                        "type": "calls_external",
                        "source": "svc-api",
                        "target": "ext-payments",
                        "label": "HTTPS 443",
                        "evidence_id": "ev-ext",
                        "metadata": {"record_id": "rec-ext"},
                    }
                ],
                "dropped_edges": [],
            }
        ),
        encoding="utf-8",
    )

    diff = build_change_diff(base_graph, new_graph)
    outputs = write_change_diff_report(tmp_path / "out", diff)

    assert diff["summary"]["base_type"] == "graph_json"
    assert diff["risks"][0]["category"] == "added_external_call"
    assert outputs["md"].exists()
    assert outputs["json"].exists()
    assert outputs["pr"].exists()
    assert "Review Required" in outputs["pr"].read_text(encoding="utf-8")


def test_change_diff_cli_writes_zero_risk_report_for_identical_clean_sample(tmp_path: Path) -> None:
    output_dir = tmp_path / "diff"

    code = main(["diff", "--base", "samples/DCP_clean_v0.1", "--new", "samples/DCP_clean_v0.1", "--output", str(output_dir)])

    assert code == 0
    payload = json.loads((output_dir / "change_diff_risk.json").read_text(encoding="utf-8"))
    assert payload["summary"]["risks"] == 0
    assert (output_dir / "pr_review_comment.md").exists()


def _append_row(ws, values: dict[str, str]) -> None:
    header_row = _header_row(ws)
    headers = [cell.value for cell in ws[header_row]]
    ws.append([values.get(str(header), "") if header else "" for header in headers])


def _remove_first_data_row(ws) -> None:
    row = _header_row(ws) + 1
    if ws.max_row >= row:
        ws.delete_rows(row)


def _header_row(ws) -> int:
    for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if "Record_ID" in row:
            return index
    raise AssertionError(f"Record_ID header row not found in {ws.title}")
